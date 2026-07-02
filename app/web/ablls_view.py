from __future__ import annotations

import datetime as dt
import os
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook


NOME_ARQUIVO_ABLLS = "planilha_grid_resultados_traduzido.xlsx"
PROTOCOLO_ABLLS = {"codigo": "ablls_r", "nome": "ABLLS-R"}
ABLLS_CODE_RE = re.compile(r"^([A-Y])(\d{1,3})$")
ABLLS_CATEGORY_RE = re.compile(r"^[A-Y]$")
API_TIMEOUT = 75
API_SYNC_TIMEOUT = 120
API_RESULTS_TIMEOUT = 120


def _texto(valor) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def _api_base_url(api_base_url: str | None = None) -> str:
    return (api_base_url or os.getenv("SKINNER_API_URL") or "http://127.0.0.1:8000").rstrip("/")


def _api_get(
    path: str,
    *,
    params: dict | None = None,
    api_base_url: str | None = None,
    timeout: int = API_TIMEOUT,
):
    resp = requests.get(
        f"{_api_base_url(api_base_url)}{path}",
        params=params,
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()


def _api_post(
    path: str,
    payload: dict,
    *,
    api_base_url: str | None = None,
    timeout: int = API_TIMEOUT,
):
    resp = requests.post(
        f"{_api_base_url(api_base_url)}{path}",
        json=payload,
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()


def localizar_arquivo_ablls(nome_arquivo: str = NOME_ARQUIVO_ABLLS) -> Path:
    base = Path(__file__).resolve().parent
    candidatos = [
        base / "data" / nome_arquivo,
        base / nome_arquivo,
        Path.cwd() / "data" / nome_arquivo,
        Path.cwd() / nome_arquivo,
        Path.home() / "Downloads" / nome_arquivo,
    ]

    for caminho in candidatos:
        if caminho.exists():
            return caminho

    return candidatos[0]


def localizar_docx_ablls() -> Path | None:
    candidatos = [
        Path(__file__).resolve().parent / "data" / "ABLLS_Corrigido_Traduzido.docx",
        Path.cwd() / "data" / "ABLLS_Corrigido_Traduzido.docx",
        Path.home() / "Downloads" / "ABLLS_Corrigido_Traduzido.docx",
    ]
    for caminho in candidatos:
        if caminho.exists():
            return caminho
    return None


def _celula_tem_borda(cell) -> bool:
    for lado in ("left", "right", "top", "bottom"):
        borda = getattr(cell.border, lado, None)
        if borda is not None and borda.style:
            return True
    return False


def _tem_borda_lado(cell, lado: str) -> bool:
    borda = getattr(cell.border, lado, None)
    return borda is not None and bool(borda.style)


def _merge_da_celula(ws, linha: int, coluna: int):
    for faixa in ws.merged_cells.ranges:
        if (
            faixa.min_row <= linha <= faixa.max_row
            and faixa.min_col <= coluna <= faixa.max_col
        ):
            return faixa
    return None


def _contar_quadradinhos(ws, linha: int, coluna_codigo: int, coluna_descricao: int) -> int:
    coluna = coluna_codigo + 1
    total = 0

    while coluna < coluna_descricao:
        faixa = _merge_da_celula(ws, linha, coluna)
        if faixa:
            coluna_final = min(faixa.max_col, coluna_descricao - 1)
            if any(_celula_tem_borda(ws.cell(linha, col)) for col in range(coluna, coluna_final + 1)):
                total += 1
            coluna = coluna_final + 1
            continue

        coluna_final = coluna
        while coluna_final < coluna_descricao - 1:
            proxima_coluna = coluna_final + 1
            if _merge_da_celula(ws, linha, proxima_coluna):
                break

            atual = ws.cell(linha, coluna_final)
            proxima = ws.cell(linha, proxima_coluna)
            tem_divisoria = _tem_borda_lado(atual, "right") or _tem_borda_lado(proxima, "left")
            if tem_divisoria:
                break

            coluna_final = proxima_coluna

        if any(_celula_tem_borda(ws.cell(linha, col)) for col in range(coluna, coluna_final + 1)):
            total += 1
        coluna = coluna_final + 1

    return max(total, 1)


@st.cache_data(show_spinner=False)
def carregar_detalhes_ablls_docx(caminho_docx: str | None) -> dict[str, dict[str, str]]:
    if not caminho_docx:
        return {}

    try:
        from docx import Document
    except ImportError:
        return {}

    path = Path(caminho_docx)
    if not path.exists():
        return {}

    doc = Document(str(path))
    detalhes: dict[str, dict[str, str]] = {}
    for table in doc.tables:
        if not table.rows:
            continue
        headers = [cell.text.strip().upper() for cell in table.rows[0].cells]
        if "ATIV." not in headers or "CRITERIO" not in headers:
            continue

        for row in table.rows[1:]:
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) < 7:
                continue
            codigo = cells[0].strip()
            if not ABLLS_CODE_RE.fullmatch(codigo):
                continue
            detalhes[codigo] = {
                "nome_atividade": cells[2],
                "objetivo_atividade": cells[3],
                "pergunta": cells[4],
                "exemplo": cells[5],
                "criterio": cells[6],
            }
    return detalhes


@st.cache_data(show_spinner=False)
def carregar_objetivos_ablls(caminho_arquivo: str, caminho_docx: str | None = None) -> pd.DataFrame:
    wb = load_workbook(caminho_arquivo, data_only=True)
    detalhes_docx = carregar_detalhes_ablls_docx(caminho_docx)
    registros = []

    for ordem_aba, nome_aba in enumerate(wb.sheetnames):
        if not nome_aba.strip().upper().startswith("ABLLS"):
            continue

        ws = wb[nome_aba]
        categorias = []

        for linha in range(1, ws.max_row + 1):
            for coluna in range(1, ws.max_column + 1):
                letra = _texto(ws.cell(linha, coluna).value)
                if not ABLLS_CATEGORY_RE.fullmatch(letra):
                    continue

                categoria = ""
                for coluna_nome in range(coluna + 1, min(ws.max_column, coluna + 7) + 1):
                    categoria = _texto(ws.cell(linha, coluna_nome).value)
                    if categoria:
                        break

                if categoria:
                    categorias.append(
                        {
                            "letra": letra,
                            "categoria": categoria,
                            "linha_categoria": linha,
                            "coluna_codigo": coluna,
                            "ordem_bloco": len(categorias),
                        }
                    )

        for categoria_info in categorias:
            letra = categoria_info["letra"]
            coluna_codigo = categoria_info["coluna_codigo"]
            coluna_descricao = coluna_codigo + 5
            if coluna_descricao > ws.max_column:
                continue

            for linha in range(1, categoria_info["linha_categoria"]):
                codigo_original = _texto(ws.cell(linha, coluna_codigo).value)
                match = ABLLS_CODE_RE.fullmatch(codigo_original)
                if not match:
                    continue

                numero = int(match.group(2))
                codigo_corrigido = f"{letra}{numero}"
                descricao = _texto(ws.cell(linha, coluna_descricao).value)
                if not descricao:
                    continue

                registros.append(
                    {
                        "aba": nome_aba.strip(),
                        "ordem_aba": ordem_aba,
                        "ordem_bloco": categoria_info["ordem_bloco"],
                        "categoria_codigo": letra,
                        "categoria": categoria_info["categoria"],
                        "codigo": codigo_corrigido,
                        "codigo_original": codigo_original if codigo_original != codigo_corrigido else "",
                        "numero": numero,
                        "descricao": descricao,
                        "max_pontos": _contar_quadradinhos(ws, linha, coluna_codigo, coluna_descricao),
                        "detalhes": detalhes_docx.get(codigo_corrigido, {}),
                    }
                )

    colunas = [
        "aba",
        "categoria_codigo",
        "categoria",
        "codigo",
        "codigo_original",
        "numero",
        "descricao",
        "max_pontos",
        "objetivo",
        "detalhes",
    ]

    if not registros:
        return pd.DataFrame(columns=colunas)

    objetivos = pd.DataFrame(registros)
    objetivos = objetivos.sort_values(["ordem_aba", "ordem_bloco", "numero"])
    objetivos = objetivos.drop_duplicates(subset=["categoria_codigo", "numero"], keep="first")
    objetivos["objetivo"] = objetivos["codigo"] + " - " + objetivos["descricao"]
    return objetivos[colunas].reset_index(drop=True)


def _sync_payload_ablls(objetivos: pd.DataFrame) -> dict:
    itens = []
    for obj in objetivos.itertuples(index=False):
        itens.append(
            {
                "codigo": obj.codigo,
                "categoria_codigo": obj.categoria_codigo,
                "categoria": obj.categoria,
                "descricao": obj.descricao,
                "max_pontos": int(obj.max_pontos),
                "aba": obj.aba,
                "numero": int(obj.numero),
                "detalhes": obj.detalhes if isinstance(obj.detalhes, dict) else {},
            }
        )
    return {"codigo": PROTOCOLO_ABLLS["codigo"], "nome": PROTOCOLO_ABLLS["nome"], "itens": itens}


def sincronizar_ablls_com_api(objetivos: pd.DataFrame, api_base_url: str | None = None) -> bool:
    cache_key = f"assessment_sync::{PROTOCOLO_ABLLS['codigo']}::{len(objetivos)}"
    if st.session_state.get(cache_key):
        return True

    try:
        protocolos = _api_get("/api/avaliacoes", api_base_url=api_base_url)
        protocolo_atual = next(
            (item for item in protocolos if item.get("code") == PROTOCOLO_ABLLS["codigo"]),
            None,
        )
        if protocolo_atual and int(protocolo_atual.get("total_itens") or 0) == len(objetivos):
            st.session_state[cache_key] = True
            return True

        _api_post(
            "/api/avaliacoes/sincronizar",
            _sync_payload_ablls(objetivos),
            api_base_url=api_base_url,
            timeout=API_SYNC_TIMEOUT,
        )
        st.session_state[cache_key] = True
        return True
    except Exception as exc:
        st.warning(f"Nao consegui sincronizar o ABLLS-R com a API: {exc}")
        return False


def _normalizar_detalhes(valor: Any) -> dict[str, Any]:
    if isinstance(valor, dict):
        return valor
    if isinstance(valor, str) and valor.strip():
        try:
            import json

            parsed = json.loads(valor)
            return parsed if isinstance(parsed, dict) else {}
        except ValueError:
            return {}
    return {}


def _score_status(pontos, max_pontos: int) -> str:
    if pontos is None or pd.isna(pontos):
        return "Nao avaliado"
    pontos = int(pontos)
    if pontos == 0:
        return "0 ponto"
    if pontos >= max_pontos:
        return "Adquirido"
    return f"{pontos}/{max_pontos}"


def _score_color(pontos, max_pontos: int) -> str:
    if pontos is None or pd.isna(pontos):
        return "#F3F4F6"
    pontos = int(pontos)
    if pontos == 0:
        return "#FEE2E2"
    if pontos >= max_pontos:
        return "#DCFCE7"
    return "#FEF3C7"


def _status_icon(pontos, max_pontos: int) -> str:
    if pontos is None or pd.isna(pontos):
        return "-"
    pontos = int(pontos)
    if pontos == 0:
        return "0"
    if pontos >= max_pontos:
        return "OK"
    return f"{pontos}/{max_pontos}"


def _render_objective_picker(
    objetivos_area: pd.DataFrame,
    *,
    protocolo_codigo: str,
    paciente_sel: str,
    data_avaliacao: dt.date,
) -> str:
    codigos = objetivos_area["codigo"].tolist()
    selected_key = f"assessment_selected::{protocolo_codigo}::{paciente_sel}::{data_avaliacao}::{objetivos_area['categoria_codigo'].iloc[0]}"
    select_key = f"select::{selected_key}"
    if selected_key not in st.session_state or st.session_state[selected_key] not in codigos:
        st.session_state[selected_key] = codigos[0]
    if select_key not in st.session_state or st.session_state[select_key] not in codigos:
        st.session_state[select_key] = st.session_state[selected_key]

    st.markdown("#### Habilidades da area")
    st.caption("Clique em uma habilidade para seleciona-la. Legenda: '-' nao avaliado, '0' zero ponto, 'OK' adquirido, 'R' tem avaliacao anterior.")

    for start in range(0, len(codigos), 6):
        cols = st.columns(6)
        for col, codigo in zip(cols, codigos[start : start + 6]):
            row = objetivos_area.loc[objetivos_area["codigo"] == codigo].iloc[0]
            pontos = None if pd.isna(row["pontos"]) else int(row["pontos"])
            max_pontos = int(row["max_pontos"])
            reavaliacao = bool(row.get("reavaliacao", False))
            selected = codigo == st.session_state[selected_key]
            label = f"{codigo} | {_status_icon(pontos, max_pontos)}"
            if reavaliacao:
                label = f"R {label}"
            if selected:
                label = f"> {label}"
            if col.button(
                label,
                key=f"pick::{selected_key}::{codigo}",
                use_container_width=True,
                help=str(row["descricao"]),
            ):
                st.session_state[selected_key] = codigo
                st.session_state[select_key] = codigo
                st.rerun()

    if st.session_state[select_key] != st.session_state[selected_key]:
        st.session_state[selected_key] = st.session_state[select_key]

    codigo_objetivo = st.selectbox(
        "Objetivo selecionado",
        options=codigos,
        index=codigos.index(st.session_state[selected_key]),
        format_func=lambda codigo: objetivos_area.loc[objetivos_area["codigo"] == codigo, "objetivo"].iloc[0],
        key=select_key,
    )
    st.session_state[selected_key] = codigo_objetivo
    return codigo_objetivo


def montar_resultados_de_api(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if df.empty:
        return df

    for coluna in ["aba", "categoria_codigo", "categoria", "codigo", "descricao", "fonte", "observacao"]:
        if coluna not in df.columns:
            df[coluna] = ""

    df["max_pontos"] = pd.to_numeric(df["max_pontos"], errors="coerce").fillna(1).astype(int)
    df["numero"] = pd.to_numeric(df.get("numero"), errors="coerce")
    df["pontos"] = pd.to_numeric(df.get("pontos"), errors="coerce")
    df["detalhes"] = df.get("detalhes", pd.Series([{}] * len(df))).map(_normalizar_detalhes)
    if "reavaliacao" not in df.columns:
        df["reavaliacao"] = False
    df["reavaliacao"] = df["reavaliacao"].fillna(False).astype(bool)
    df["pontos_anteriores"] = pd.to_numeric(df.get("pontos_anteriores"), errors="coerce")
    df["pontuou"] = df["pontos"].map(lambda valor: pd.notna(valor) and int(valor) > 0)
    df["adquirido"] = df.apply(
        lambda row: pd.notna(row["pontos"]) and int(row["pontos"]) >= int(row["max_pontos"]),
        axis=1,
    )
    df["status"] = df.apply(
        lambda row: _score_status(row["pontos"], int(row["max_pontos"])),
        axis=1,
    )
    df["tipo_avaliacao"] = df["reavaliacao"].map(lambda valor: "Reavaliacao" if valor else "Inicial")
    df["objetivo"] = df["codigo"].astype(str) + " - " + df["descricao"].astype(str)
    return df.sort_values(["categoria_codigo", "numero", "codigo"], na_position="last").reset_index(drop=True)


def _nome_arquivo_seguro(valor: str) -> str:
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", str(valor)).strip("_")
    return texto or "paciente"


def _texto_pdf(canvas, texto: str, x: float, y: float, largura: float, fonte: str, tamanho: float):
    from reportlab.pdfbase.pdfmetrics import stringWidth

    texto = _texto(texto)
    if stringWidth(texto, fonte, tamanho) <= largura:
        canvas.drawString(x, y, texto)
        return

    while texto and stringWidth(texto + "...", fonte, tamanho) > largura:
        texto = texto[:-1]
    canvas.drawString(x, y, texto + "...")


def gerar_pdf_cascata_ablls(
    paciente_sel: str,
    protocolo_nome: str,
    data_avaliacao: dt.date,
    resultados_df: pd.DataFrame,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas

    buffer = BytesIO()
    page_w, page_h = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    margem = 24
    gutter = 12
    largura_coluna = (page_w - (2 * margem) - (2 * gutter)) / 3
    topo = page_h - 58
    base = 34
    altura_disponivel = topo - base

    for aba, df_aba in resultados_df.groupby("aba", sort=False, dropna=False):
        categorias = list(df_aba.groupby(["categoria_codigo", "categoria"], sort=False, dropna=False))
        for inicio in range(0, len(categorias), 3):
            grupo_categorias = categorias[inicio : inicio + 3]

            pdf.setFillColor(colors.HexColor("#1F2937"))
            pdf.setFont("Helvetica-Bold", 13)
            pdf.drawString(margem, page_h - 28, f"{protocolo_nome} - Cascata de habilidades")
            pdf.setFont("Helvetica", 9)
            pdf.drawString(
                margem,
                page_h - 43,
                f"Paciente: {paciente_sel} | Data: {data_avaliacao.strftime('%d/%m/%Y')} | Aba: {_texto(aba) or '-'}",
            )

            for indice, ((categoria_codigo, categoria), df_cat) in enumerate(grupo_categorias):
                x = margem + indice * (largura_coluna + gutter)
                y_titulo = topo + 12
                df_cat = df_cat.sort_values("numero", ascending=False, na_position="last")
                row_h = max(6.2, min(11, altura_disponivel / max(len(df_cat), 1)))
                fonte_row = max(4.6, min(7.2, row_h - 2.2))
                square = max(4.4, min(7.0, row_h - 1.4))

                pdf.setFillColor(colors.HexColor("#0F766E"))
                pdf.roundRect(x, y_titulo, largura_coluna, 13, 2, stroke=0, fill=1)
                pdf.setFillColor(colors.white)
                pdf.setFont("Helvetica-Bold", 7.5)
                _texto_pdf(
                    pdf,
                    f"{categoria_codigo} - {categoria}",
                    x + 4,
                    y_titulo + 3.5,
                    largura_coluna - 8,
                    "Helvetica-Bold",
                    7.5,
                )

                for pos, obj in enumerate(df_cat.itertuples(index=False)):
                    y = topo - ((pos + 1) * row_h)
                    pontos = None if pd.isna(obj.pontos) else int(obj.pontos)
                    max_pontos = int(obj.max_pontos)
                    reavaliacao = bool(getattr(obj, "reavaliacao", False))

                    pdf.setFillColor(colors.HexColor(_score_color(pontos, max_pontos)))
                    pdf.setStrokeColor(colors.HexColor("#2563EB" if reavaliacao else "#E5E7EB"))
                    pdf.setLineWidth(1.2 if reavaliacao else 0.5)
                    pdf.rect(x, y, largura_coluna, row_h, stroke=1, fill=1)
                    pdf.setLineWidth(0.5)

                    pdf.setFillColor(colors.HexColor("#111827"))
                    pdf.setFont("Helvetica-Bold", fonte_row)
                    pdf.drawString(x + 2, y + 1.6, f"R {obj.codigo}" if reavaliacao else str(obj.codigo))

                    sx = x + 22
                    sy = y + (row_h - square) / 2
                    for ponto in range(max_pontos):
                        preenchido = pontos is not None and ponto < pontos
                        pdf.setFillColor(colors.HexColor("#22C55E") if preenchido else colors.white)
                        pdf.setStrokeColor(colors.HexColor("#6B7280"))
                        pdf.rect(sx + ponto * (square + 1.2), sy, square, square, stroke=1, fill=1)

                    texto_x = sx + max_pontos * (square + 1.2) + 4
                    pdf.setFillColor(colors.HexColor("#374151"))
                    pdf.setFont("Helvetica", fonte_row)
                    _texto_pdf(
                        pdf,
                        obj.descricao,
                        texto_x,
                        y + 1.6,
                        x + largura_coluna - texto_x - 2,
                        "Helvetica",
                        fonte_row,
                    )

            pdf.setFillColor(colors.HexColor("#6B7280"))
            pdf.setFont("Helvetica", 7)
            pdf.drawRightString(
                page_w - margem,
                16,
                "Verde = adquirido | amarelo = parcial | vermelho = 0 | cinza = nao avaliado | borda azul = reavaliacao",
            )
            pdf.showPage()

    pdf.save()
    buffer.seek(0)
    return buffer.read()


def _render_detalhes_objetivo(detalhes: dict[str, Any]):
    if not detalhes:
        return

    with st.expander("Criterio do protocolo traduzido"):
        campos = [
            ("Objetivo da atividade", "objetivo_atividade"),
            ("Pergunta avaliativa", "pergunta"),
            ("Exemplo", "exemplo"),
            ("Criterio de pontuacao", "criterio"),
        ]
        for titulo, chave in campos:
            if detalhes.get(chave):
                st.markdown(f"**{titulo}**")
                st.write(detalhes[chave])


def _render_vinculo_bhave(protocolo_codigo: str, objetivo: pd.Series, api_base_url: str | None):
    with st.expander("Vincular este item com a biblioteca bHave"):
        try:
            biblioteca = _api_get("/api/avaliacoes/biblioteca_bhave", api_base_url=api_base_url)
            vinculos = _api_get(
                f"/api/avaliacoes/{protocolo_codigo}/vinculos_bhave",
                params={"codigo_item": objetivo["codigo"]},
                api_base_url=api_base_url,
            )
        except Exception as exc:
            st.caption(f"Biblioteca bHave indisponivel: {exc}")
            return

        if vinculos:
            st.caption("Vinculos atuais")
            st.dataframe(pd.DataFrame(vinculos), hide_index=True, use_container_width=True)

        nomes = [item["name"] for item in biblioteca if item.get("name")]
        if not nomes:
            st.caption("Nenhum programa encontrado na biblioteca bHave.")
            return

        programa = st.selectbox(
            "Programa bHave que comprova esta habilidade",
            options=nomes,
            key=f"link_program::{protocolo_codigo}::{objetivo['codigo']}",
        )
        pontos_auto = st.number_input(
            "Pontos automaticos quando o criterio for atingido",
            min_value=0,
            max_value=int(objetivo["max_pontos"]),
            value=int(objetivo["max_pontos"]),
            step=1,
            key=f"link_points::{protocolo_codigo}::{objetivo['codigo']}",
        )

        escolhido = next((item for item in biblioteca if item.get("name") == programa), None)
        if escolhido:
            threshold = escolhido.get("mastery_threshold_percent") or 90
            dias = escolhido.get("mastery_days") or 3
            st.caption(f"Criterio bHave: {threshold}% por {dias} sessoes/dias configurados na biblioteca.")

        if st.button("Salvar vinculo bHave", use_container_width=True):
            payload = {
                "codigo_item": objetivo["codigo"],
                "programa_biblioteca": programa,
                "pontos_automaticos": int(pontos_auto),
            }
            try:
                _api_post(f"/api/avaliacoes/{protocolo_codigo}/vincular_bhave", payload, api_base_url=api_base_url)
                st.success("Vinculo salvo.")
            except Exception as exc:
                st.error(f"Erro ao salvar vinculo: {exc}")


def _aplicar_sugestoes_bhave(
    protocolo_codigo: str,
    paciente_sel: str,
    data_avaliacao: dt.date,
    resultados_df: pd.DataFrame,
    api_base_url: str | None,
) -> int:
    sugestoes = _api_get(
        f"/api/avaliacoes/{protocolo_codigo}/sugestoes_bhave",
        params={"paciente": paciente_sel},
        api_base_url=api_base_url,
    )
    pontos_atuais = {
        row.codigo: None if pd.isna(row.pontos) else int(row.pontos)
        for row in resultados_df.itertuples(index=False)
    }
    aplicadas = 0
    for sugestao in sugestoes:
        sugerido = int(sugestao["pontos_sugeridos"])
        atual = pontos_atuais.get(sugestao["codigo"])
        if atual is not None and atual >= sugerido:
            continue

        payload = {
            "paciente": paciente_sel,
            "codigo_item": sugestao["codigo"],
            "pontos": sugerido,
            "data_avaliacao": data_avaliacao.isoformat(),
            "observacao": f"Pontuado automaticamente pelo bHave: {sugestao['programa']}",
            "fonte": "auto_bhave",
        }
        _api_post(f"/api/avaliacoes/{protocolo_codigo}/pontuar", payload, api_base_url=api_base_url)
        aplicadas += 1
    return aplicadas


def _styled_area_table(df: pd.DataFrame):
    tabela = df[[
        "codigo",
        "descricao",
        "pontos",
        "pontos_anteriores",
        "max_pontos",
        "status",
        "tipo_avaliacao",
        "fonte",
    ]].copy()
    tabela = tabela.rename(
        columns={
            "codigo": "Codigo",
            "descricao": "Descricao",
            "pontos": "Pontos",
            "pontos_anteriores": "Pontuacao anterior",
            "max_pontos": "Maximo",
            "status": "Status",
            "tipo_avaliacao": "Tipo",
            "fonte": "Fonte",
        }
    )

    def color_row(row):
        if row["Tipo"] == "Reavaliacao":
            return ["background-color: #DBEAFE; color: #1E3A8A"] * len(row)
        return [""] * len(row)

    return tabela.style.apply(color_row, axis=1)


def render_ablls_module(
    paciente_sel,
    df_p_raw=None,
    caminho_arquivo=None,
    caminho_docx=None,
    api_base_url=None,
):
    st.subheader(f"Avaliacoes: {paciente_sel}")

    caminho = Path(caminho_arquivo) if caminho_arquivo else localizar_arquivo_ablls()
    docx_path = Path(caminho_docx) if caminho_docx else localizar_docx_ablls()
    if not caminho.exists():
        st.error(f"Arquivo ABLLS-R nao encontrado: {caminho}")
        return pd.DataFrame()

    try:
        objetivos = carregar_objetivos_ablls(str(caminho), str(docx_path) if docx_path else None)
    except Exception as exc:
        st.error(f"Erro ao ler a planilha ABLLS-R: {exc}")
        return pd.DataFrame()

    if objetivos.empty:
        st.warning("Nenhum objetivo ABLLS-R foi encontrado.")
        return pd.DataFrame()

    api_ok = sincronizar_ablls_com_api(objetivos, api_base_url=api_base_url)
    if not api_ok:
        st.error("A avaliacao precisa da API ligada para salvar pontuacoes no Supabase.")
        return pd.DataFrame()

    try:
        protocolos_api = _api_get("/api/avaliacoes", api_base_url=api_base_url)
        protocolos = [{"codigo": item["code"], "nome": item["name"]} for item in protocolos_api]
    except Exception as exc:
        st.error(f"Nao consegui carregar os protocolos avaliativos: {exc}")
        return pd.DataFrame()

    if not protocolos:
        protocolos = [PROTOCOLO_ABLLS]

    protocolo_codigo = st.selectbox(
        "Tipo de avaliacao",
        options=[p["codigo"] for p in protocolos],
        format_func=lambda codigo: next((p["nome"] for p in protocolos if p["codigo"] == codigo), codigo),
    )
    protocolo_nome = next((p["nome"] for p in protocolos if p["codigo"] == protocolo_codigo), protocolo_codigo)
    data_avaliacao = st.date_input("Data da avaliacao", value=dt.date.today())

    try:
        rows = _api_get(
            f"/api/avaliacoes/{protocolo_codigo}/resultados",
            params={"paciente": paciente_sel, "data_avaliacao": data_avaliacao.isoformat()},
            api_base_url=api_base_url,
            timeout=API_RESULTS_TIMEOUT,
        )
        resultados_df = montar_resultados_de_api(rows)
    except Exception as exc:
        st.error(f"Nao consegui buscar resultados salvos na API: {exc}")
        return pd.DataFrame()

    if resultados_df.empty:
        st.warning("Este protocolo ainda nao tem itens cadastrados.")
        return resultados_df

    pontos_obtidos = int(resultados_df["pontos"].fillna(0).sum())
    pontos_possiveis = int(resultados_df["max_pontos"].sum())
    avaliados = int(resultados_df["pontos"].notna().sum())
    completos = int(resultados_df["adquirido"].sum())
    itens_com_historico = int(resultados_df["reavaliacao"].sum()) if "reavaliacao" in resultados_df.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Itens", len(resultados_df))
    c2.metric("Avaliados", avaliados)
    c3.metric("Adquiridos", completos)
    c4.metric("Pontos", f"{pontos_obtidos}/{pontos_possiveis}")
    st.progress(pontos_obtidos / pontos_possiveis if pontos_possiveis else 0)
    if itens_com_historico:
        st.info(
            f"Esta data tem {itens_com_historico} item(ns) com pontuacao em avaliacao anterior. "
            "Eles aparecem marcados com R e em azul na tabela da area."
        )

    col_auto, col_botao = st.columns([2, 1])
    auto_bhave = col_auto.checkbox(
        "Pontuar automaticamente pelo bHave quando o criterio estiver batido",
        value=False,
    )
    aplicar_bhave_agora = col_botao.button("Aplicar bHave agora", use_container_width=True)

    if auto_bhave or aplicar_bhave_agora:
        try:
            aplicadas = _aplicar_sugestoes_bhave(
                protocolo_codigo,
                str(paciente_sel),
                data_avaliacao,
                resultados_df,
                api_base_url,
            )
            if aplicadas:
                st.success(f"{aplicadas} habilidade(s) pontuada(s) pelo bHave.")
                st.rerun()
            elif aplicar_bhave_agora:
                st.info("Nenhuma habilidade vinculada atingiu criterio de aquisicao agora.")
        except Exception as exc:
            st.error(f"Erro ao aplicar sugestoes bHave: {exc}")

    categorias = resultados_df[["categoria_codigo", "categoria"]].drop_duplicates()
    categoria_codigo = st.selectbox(
        "Area",
        options=categorias["categoria_codigo"].tolist(),
        format_func=lambda codigo: (
            f"{codigo} - "
            f"{categorias.loc[categorias['categoria_codigo'] == codigo, 'categoria'].iloc[0]}"
        ),
    )

    objetivos_area = resultados_df[resultados_df["categoria_codigo"] == categoria_codigo].copy()
    codigo_objetivo = _render_objective_picker(
        objetivos_area,
        protocolo_codigo=protocolo_codigo,
        paciente_sel=str(paciente_sel),
        data_avaliacao=data_avaliacao,
    )

    objetivo = objetivos_area.loc[objetivos_area["codigo"] == codigo_objetivo].iloc[0]
    max_pontos = int(objetivo["max_pontos"])
    pontos_salvos = None if pd.isna(objetivo["pontos"]) else int(objetivo["pontos"])
    opcoes = [None] + list(range(max_pontos + 1))

    st.markdown("#### Pontuar habilidade")
    st.markdown(f"**{objetivo['codigo']} - {_score_status(pontos_salvos, max_pontos)}**")
    st.write(str(objetivo["descricao"]))
    st.caption(f"{objetivo['aba']} / {objetivo['categoria']} / maximo {max_pontos} pontos")
    if bool(objetivo.get("reavaliacao", False)):
        anterior = objetivo.get("pontos_anteriores")
        texto_anterior = "-" if pd.isna(anterior) else str(int(anterior))
        st.info(f"Reavaliacao: este item tinha pontuacao anterior de {texto_anterior}.")

    detalhes = objetivo.get("detalhes") or {}
    _render_detalhes_objetivo(detalhes if isinstance(detalhes, dict) else {})

    pontos_key = f"assessment_score::{protocolo_codigo}::{paciente_sel}::{data_avaliacao}::{codigo_objetivo}"
    novo_ponto = st.radio(
        "Pontuacao",
        options=opcoes,
        index=opcoes.index(pontos_salvos),
        horizontal=True,
        key=pontos_key,
        format_func=lambda valor: "Nao avaliado" if valor is None else f"{valor} ponto{'s' if valor != 1 else ''}",
    )

    if novo_ponto != pontos_salvos:
        payload = {
            "paciente": str(paciente_sel),
            "codigo_item": codigo_objetivo,
            "pontos": novo_ponto,
            "data_avaliacao": data_avaliacao.isoformat(),
            "observacao": None,
            "fonte": "manual",
        }
        try:
            _api_post(f"/api/avaliacoes/{protocolo_codigo}/pontuar", payload, api_base_url=api_base_url)
            st.rerun()
        except Exception as exc:
            st.error(f"Erro ao salvar pontuacao: {exc}")

    _render_vinculo_bhave(protocolo_codigo, objetivo, api_base_url)

    with st.expander("Objetivos desta area", expanded=True):
        st.dataframe(
            _styled_area_table(objetivos_area),
            hide_index=True,
            use_container_width=True,
            height=360,
        )

    nome_seguro = _nome_arquivo_seguro(str(paciente_sel))
    try:
        pdf_bytes = gerar_pdf_cascata_ablls(str(paciente_sel), protocolo_nome, data_avaliacao, resultados_df)
        st.download_button(
            "Baixar PDF da cascata pintada",
            data=pdf_bytes,
            file_name=f"{protocolo_codigo}_cascata_{nome_seguro}_{data_avaliacao.isoformat()}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    except ImportError:
        st.warning("Instale `reportlab` no ambiente do Streamlit para gerar o PDF.")

    csv_bytes = resultados_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "Baixar resultados salvos (.csv)",
        data=csv_bytes,
        file_name=f"{protocolo_codigo}_resultados_{nome_seguro}_{data_avaliacao.isoformat()}.csv",
        mime="text/csv",
        use_container_width=True,
    )

    with st.expander("Resultados salvos"):
        st.dataframe(resultados_df, hide_index=True, use_container_width=True, height=420)

    return resultados_df
